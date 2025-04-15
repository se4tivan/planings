import os
import sys
import requests
import schedule
import time
from datetime import datetime
import pytz
from openpyxl import load_workbook
from telegram.ext import Updater, CommandHandler
from flask import Flask
from io import BytesIO
import logging

# Настройка логирования
logging.basicConfig(
    format='%(asctime)s - %(levelname)s - %(message)s',
    level=logging.INFO,
    handlers=[
        logging.FileHandler('bot.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# Создаем Flask приложение
app = Flask(__name__)

@app.route('/')
def home():
    return "Bot is running!"

# Константы
TELEGRAM_BOT_TOKEN = "7738985245:AAEtSoP8Jc8vRuNEZd5RJQpJ-0NCMaW7Xys"
CHAT_ID = "-1002017911073"  # ID чата для уведомлений
MOSCOW_TZ = pytz.timezone('Europe/Moscow')
EXCEL_URL = "https://cloud.toplubricants.ru/s/yWSpybsLPZkjjzp/download"

# Хранение предыдущих данных для сравнения
previous_data = {}

def is_workday():
    """Проверка рабочего времени (Пн-Пт, 8:00-18:00)"""
    try:
        now = datetime.now(MOSCOW_TZ)
        return (
            now.weekday() < 5 and  # Понедельник = 0, Воскресенье = 6
            8 <= now.hour < 18
        )
    except Exception as e:
        logger.error(f"Ошибка при проверке рабочего времени: {str(e)}")
        return False

def download_excel():
    """Загрузка Excel файла по ссылке"""
    try:
        logger.info("Начало загрузки Excel файла")
        response = requests.get(EXCEL_URL)
        response.raise_for_status()
        logger.info("Excel файл успешно загружен")
        return response.content
    except Exception as e:
        logger.error(f"Ошибка при загрузке Excel файла: {str(e)}")
        return None

def read_planning_data():
    """Чтение данных из Excel файла"""
    try:
        excel_content = download_excel()
        if excel_content is None:
            return None
        
        logger.info("Начало чтения данных из Excel")
        wb = load_workbook(filename=BytesIO(excel_content), read_only=True)
        ws = wb.active
        current_data = {}
        
        # Получаем все строки
        rows = list(ws.iter_rows(min_row=2))
        logger.info(f"Всего строк в файле: {len(rows)}")
        
        for row_num, row in enumerate(rows, start=2):  # Начинаем с 2, так как пропускаем заголовок
            try:
                # Проверяем, что строка содержит достаточно столбцов
                if len(row) < 13:
                    continue
                    
                order_num = row[11].value  # Столбец L
                driver = row[12].value  # Столбец M
                vehicle = row[8].value  # Столбец I
                
                # Пропускаем пустые строки
                if not order_num or not driver or not vehicle:
                    continue
                
                # Логируем каждую строку для отладки
                logger.info(f"Строка {row_num}: номер={order_num}, водитель={driver}, машина={vehicle}")
                
                current_data[str(order_num)] = {
                    'date': row[0].value,  # Столбец A
                    'driver': driver,
                    'vehicle': vehicle,
                    'row_num': row_num  # Добавляем номер строки
                }
            except Exception as e:
                logger.error(f"Ошибка при обработке строки {row_num}: {str(e)}")
                continue
        
        logger.info(f"Успешно прочитано {len(current_data)} заказов")
        return current_data
    except Exception as e:
        logger.error(f"Ошибка при чтении Excel файла: {str(e)}")
        return None

def check_changes():
    """Проверка изменений в планинге"""
    try:
        if not is_workday():
            return
        
        global previous_data
        current_data = read_planning_data()
        
        if current_data is None:
            return
        
        if not previous_data:
            previous_data = current_data
            return
        
        messages = []
        
        # Проверка новых и измененных заказов
        for order_num, current_info in current_data.items():
            # Пропускаем строки до 744
            if current_info['row_num'] < 744:
                continue
                
            if order_num not in previous_data:
                messages.append(
                    f"📅 {current_info['date'].strftime('%d.%m.%Y')}\n"
                    f"🆕 Новый заказ: {order_num}\n"
                    f"Водитель: {current_info['driver']}\n"
                    f"Транспорт: {current_info['vehicle']}"
                )
            else:
                prev_info = previous_data[order_num]
                if current_info['driver'] != prev_info['driver']:
                    messages.append(
                        f"👨‍✈️ Замена водителя в заказе {order_num}\n"
                        f"Было: {prev_info['driver']}\n"
                        f"Стало: {current_info['driver']}"
                    )
                if current_info['vehicle'] != prev_info['vehicle']:
                    messages.append(
                        f"🚛 Замена транспорта в заказе {order_num}\n"
                        f"Было: {prev_info['vehicle']}\n"
                        f"Стало: {current_info['vehicle']}"
                    )
        
        # Проверка удаленных заказов
        for order_num, prev_info in previous_data.items():
            # Пропускаем строки до 744
            if prev_info['row_num'] < 744:
                continue
                
            if order_num not in current_data:
                messages.append(f"❌ Удален заказ: {order_num}")
        
        # Отправка сообщений в Telegram
        if messages:
            send_telegram_messages(messages)
        
        previous_data = current_data
    except Exception as e:
        logger.error(f"Ошибка при проверке изменений: {str(e)}")

def send_telegram_messages(messages):
    """Отправка сообщений в Telegram канал"""
    try:
        updater = Updater(TELEGRAM_BOT_TOKEN)
        for message in messages:
            try:
                updater.bot.send_message(
                    chat_id=CHAT_ID,
                    text=message,
                    parse_mode='HTML'
                )
                logger.info(f"Сообщение успешно отправлено в канал: {message[:50]}...")
            except Exception as e:
                logger.error(f"Ошибка при отправке сообщения в канал: {str(e)}")
    except Exception as e:
        logger.error(f"Критическая ошибка при отправке сообщений: {str(e)}")

def start_monitoring():
    """Запуск процесса мониторинга"""
    try:
        # Проверка каждые 5 минут в рабочее время
        schedule.every(5).minutes.do(check_changes)
        
        while True:
            schedule.run_pending()
            time.sleep(60)
    except:
        pass

def format_phone_number(phone):
    """Форматирование номера телефона в формат 8(888)888-88-88"""
    try:
        # Удаляем все нецифровые символы
        digits = ''.join(filter(str.isdigit, str(phone)))
        
        # Проверяем длину номера
        if len(digits) != 11:
            return phone  # Возвращаем оригинальный номер, если формат не подходит
        
        # Форматируем номер
        return f"8({digits[1:4]}){digits[4:7]}-{digits[7:9]}-{digits[9:]}"
    except:
        return phone  # Возвращаем оригинальный номер в случае ошибки

def status(update, context):
    """Обработчик команды /status"""
    try:
        chat_id = update.effective_chat.id
        logger.info(f"Получена команда /status от пользователя {chat_id}")
        
        if is_workday():
            response = 'Бот активен и отслеживает изменения в планинге.'
        else:
            response = 'Бот неактивен: нерабочее время или выходной день.'
        
        update.message.reply_text(response)
        logger.info(f"Отправлен статус пользователю {chat_id}: {response}")
    except Exception as e:
        logger.error(f"Ошибка при обработке команды /status: {str(e)}")
        try:
            update.message.reply_text("Произошла ошибка при проверке статуса")
        except:
            pass

def chatid(update, context):
    """Обработчик команды /chatid"""
    try:
        chat_id = update.effective_chat.id
        chat_type = update.effective_chat.type
        update.message.reply_text(f'ID этого чата: {chat_id}\nТип чата: {chat_type}')
    except Exception as e:
        logger.error(f"Ошибка при обработке команды /chatid: {str(e)}")

def start(update, context):
    """Обработчик команды /start"""
    try:
        chat_id = update.effective_chat.id
        logger.info(f"Получена команда /start от пользователя {chat_id}")
        update.message.reply_text('Бот запущен и мониторит планинг!')
    except Exception as e:
        logger.error(f"Ошибка при обработке команды /start: {str(e)}")

def send_test_message():
    """Отправка тестового сообщения при запуске бота"""
    try:
        updater = Updater(TELEGRAM_BOT_TOKEN)
        message = (
            "🤖 Бот успешно запущен!\n\n"
            "Доступные команды:\n"
            "/start - запуск бота\n"
            "/status - проверка статуса\n"
            "/chatid - получение ID чата\n\n"
            "Бот будет отслеживать изменения в планинге и отправлять уведомления."
        )
        updater.bot.send_message(chat_id=CHAT_ID, text=message)
        logger.info("Тестовое сообщение успешно отправлено")
    except Exception as e:
        logger.error(f"Ошибка при отправке тестового сообщения: {str(e)}")

def check_message(update, context):
    """Обработчик текстового сообщения 'Проверка'"""
    try:
        chat_id = update.effective_chat.id
        message_text = update.message.text.lower()
        
        if message_text == "проверка":
            logger.info(f"Получено сообщение 'Проверка' от пользователя {chat_id}")
            
            # Отправляем сообщение в канал
            updater = Updater(TELEGRAM_BOT_TOKEN)
            status_message = (
                "🤖 Бот активен!\n\n"
                "Статус:\n"
                f"• Канал: {CHAT_ID}\n"
                f"• Рабочее время: {'Да' if is_workday() else 'Нет'}\n"
                "• Мониторинг изменений: Активен"
            )
            
            try:
                updater.bot.send_message(
                    chat_id=CHAT_ID,
                    text=status_message,
                    parse_mode='HTML'
                )
                logger.info("Статус бота отправлен в канал")
            except Exception as e:
                logger.error(f"Ошибка при отправке статуса в канал: {str(e)}")
                update.message.reply_text("Ошибка при отправке статуса в канал")
    except Exception as e:
        logger.error(f"Ошибка при обработке текстового сообщения: {str(e)}")
        try:
            update.message.reply_text("Произошла ошибка при обработке команды")
        except:
            pass

def main():
    try:
        # Настройка бота
        updater = Updater(TELEGRAM_BOT_TOKEN)
        dp = updater.dispatcher
        
        # Добавление обработчиков команд
        dp.add_handler(CommandHandler("start", start))
        dp.add_handler(CommandHandler("status", status))
        dp.add_handler(CommandHandler("chatid", chatid))
        
        # Добавление обработчика текстовых сообщений
        from telegram.ext import MessageHandler, Filters
        dp.add_handler(MessageHandler(Filters.text & ~Filters.command, check_message))
        
        # Запуск бота
        updater.start_polling()
        logger.info("Бот успешно запущен")
        
        # Отправка тестового сообщения
        send_test_message()
        
        # Запуск мониторинга в отдельном потоке
        import threading
        monitoring_thread = threading.Thread(target=start_monitoring)
        monitoring_thread.daemon = True
        monitoring_thread.start()
        logger.info("Мониторинг запущен в отдельном потоке")
        
        # Запуск Flask сервера
        port = int(os.environ.get("PORT", 8080))
        app.run(host='0.0.0.0', port=port)
        
    except Exception as e:
        logger.error(f"Критическая ошибка при запуске бота: {str(e)}")
        sys.exit(1)

if __name__ == '__main__':
    try:
        main()
    except:
        sys.exit(1) 
